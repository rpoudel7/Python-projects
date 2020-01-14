import openpyxl 
from openpyxl.chart import BarChart, Reference 


# def process_workbook(inputFile):

wb = openpyxl.load_workbook("Fees.xlsx")
sheet = wb["Sheet1"]
#cell = x["a1"]
#cell = sheet.cell(1,1)
#print(sheet.max_row)

for row in range (2, sheet.max_row +1):
	#print(row) 
	cell = sheet.cell (row,3)
	corrected_value = cell.value *1.1
	print(corrected_value)
	# crate a new column to save a new corrected prices and give them the value 
	new_price = sheet.cell(row,4)
	new_price.value = corrected_value  
  
# select the values from 4th column 
values = Reference(sheet, min_row =2, max_row = sheet.max_row, min_col=4, max_col =4)
chart = BarChart()
chart.add_data(values)

sheet.add_chart(chart,"b10")  # e2 is where the chart is added 
wb.save("Fees.xlsx")  # saves  in a new file 


